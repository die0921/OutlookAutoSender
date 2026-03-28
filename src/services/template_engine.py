import re
import openpyxl
from typing import Dict, Tuple, List
from jinja2 import Environment, BaseLoader, Undefined
from src.models.template import Template, BodyPart
from src.data.template_loader import TemplateLoader


class TemplateEngine:
    """模板渲染引擎：变量替换 + Excel 区域嵌入"""

    def __init__(self, template_loader: TemplateLoader):
        self.template_loader = template_loader
        self._jinja_env = Environment(
            loader=BaseLoader(),
            variable_start_string='{{',
            variable_end_string='}}',
            undefined=Undefined,  # silently ignore missing vars
        )

    def render(self, template_name: str, variables: Dict[str, str]) -> Tuple[str, str]:
        """渲染模板，返回 (subject, body)

        Args:
            template_name: 模板名称
            variables: 变量字典（Excel 行数据）

        Returns:
            (rendered_subject, rendered_body)
        """
        template = self.template_loader.load(template_name)
        subject = self._render_text(template.subject, variables)
        body = self._render_body(template, variables)
        return subject, body

    def get_template_variables(self, template_name: str) -> List[str]:
        """提取模板中使用的变量名列表"""
        template = self.template_loader.load(template_name)
        variables: set = set()

        # 从 subject 提取
        variables.update(self._extract_variables(template.subject))

        # 从 body_parts 提取
        for part in template.body_parts:
            if part.type == 'text' and part.content:
                variables.update(self._extract_variables(part.content))
            if part.source:
                variables.update(self._extract_variables(part.source))

        # 从 data_sources 文件路径提取
        for ds in template.data_sources.values():
            variables.update(self._extract_variables(ds.file))

        return sorted(variables)

    def _render_text(self, text: str, variables: Dict[str, str]) -> str:
        """用 Jinja2 渲染文本中的变量"""
        try:
            tmpl = self._jinja_env.from_string(text)
            return tmpl.render(**variables)
        except Exception:
            return text

    def _render_body(self, template: Template, variables: Dict[str, str]) -> str:
        """渲染模板正文（组合多个 body_parts）"""
        parts = []
        for part in template.body_parts:
            if part.type == 'text':
                rendered = self._render_text(part.content or '', variables)
                parts.append(rendered)
            elif part.type == 'excel_range':
                rendered = self._render_excel_range(part, variables)
                parts.append(rendered)

        separator = '\n' if template.body_type == 'text' else ''
        body = separator.join(parts)

        if template.body_type == 'html':
            return f'<html><body>{body}</body></html>'
        return body

    def _render_excel_range(self, part: BodyPart, variables: Dict[str, str]) -> str:
        """从 Excel 文件读取区域数据并渲染为表格"""
        if not part.source:
            return ''

        # The file path may be in variables or directly specified
        file_path = variables.get(part.source, part.source)
        file_path = self._render_text(file_path, variables)

        if not file_path:
            return ''

        try:
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            ws = wb[part.sheet] if part.sheet else wb.active
            rows = self._read_range(ws, part.range)
            wb.close()

            if part.format == 'table':
                return self._rows_to_html_table(rows, part.include_header or False)
            else:
                return self._rows_to_text(rows)
        except Exception:
            return f'[无法读取数据: {file_path}]'

    def _read_range(self, ws, range_str) -> List[List]:
        """读取工作表区域数据"""
        if range_str:
            rows = []
            for row in ws[range_str]:
                rows.append([cell.value for cell in row])
            return rows
        else:
            return [[cell.value for cell in row] for row in ws.iter_rows()]

    def _rows_to_html_table(self, rows: List[List], include_header: bool) -> str:
        """将行数据转换为 HTML 表格"""
        if not rows:
            return '<table></table>'
        html = '<table border="1">'
        start = 0
        if include_header and rows:
            html += '<tr>'
            for cell in rows[0]:
                html += f'<th>{cell if cell is not None else ""}</th>'
            html += '</tr>'
            start = 1
        for row in rows[start:]:
            html += '<tr>'
            for cell in row:
                html += f'<td>{cell if cell is not None else ""}</td>'
            html += '</tr>'
        html += '</table>'
        return html

    def _rows_to_text(self, rows: List[List]) -> str:
        """将行数据转换为纯文本"""
        lines = []
        for row in rows:
            lines.append('\t'.join(str(c) if c is not None else '' for c in row))
        return '\n'.join(lines)

    def _extract_variables(self, text: str) -> List[str]:
        """从模板文本中提取变量名"""
        return re.findall(r'\{\{(\w+)\}\}', text)
