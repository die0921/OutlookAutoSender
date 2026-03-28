import openpyxl
from datetime import datetime, date
from typing import List, Optional, Dict, Any
from src.models.email_task import EmailTask
from src.models.config import ExcelConfig


class ExcelReader:
    """读取 Excel 文件，转换为 EmailTask 列表"""

    def __init__(self, config: ExcelConfig):
        self.config = config

    def read_tasks(self, file_path: str) -> List[EmailTask]:
        """读取 Excel 文件，返回 EmailTask 列表"""
        wb = openpyxl.load_workbook(file_path)
        ws = wb[self.config.sheet_name]

        # 读取表头行，建立列名到列索引的映射
        headers = {}
        for col_idx, cell in enumerate(ws[1], start=1):
            if cell.value:
                headers[str(cell.value)] = col_idx

        tasks = []
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            # 跳过空行（收件人为空）
            recipient_col = headers.get(self.config.columns.get('recipient', '收件人'))
            if not recipient_col or not row[recipient_col - 1]:
                continue

            task = self._row_to_task(row_idx, row, headers)
            tasks.append(task)

        wb.close()
        return tasks

    def _row_to_task(self, row_idx: int, row: tuple, headers: Dict[str, int]) -> EmailTask:
        """将 Excel 行转换为 EmailTask"""
        cols = self.config.columns

        def get_val(col_key: str) -> Any:
            col_name = cols.get(col_key, '')
            col_idx = headers.get(col_name)
            if col_idx is None:
                return None
            val = row[col_idx - 1]
            return val

        # 提取已知列名之外的列作为变量
        known_cols = set(cols.values())
        variables = {}
        for header_name, col_idx in headers.items():
            if header_name not in known_cols:
                val = row[col_idx - 1]
                if val is not None:
                    variables[header_name] = val

        return EmailTask(
            row_index=row_idx,
            recipient=str(get_val('recipient') or ''),
            cc=str(get_val('cc') or ''),
            subject=str(get_val('subject') or ''),
            attachments=str(get_val('attachments') or ''),
            status=str(get_val('status') or '待发送'),
            send_mode=str(get_val('send_mode') or '一次性'),
            template_name=str(get_val('template') or ''),
            schedule_type=str(get_val('schedule_type') or ''),
            schedule_params=str(get_val('schedule_params') or ''),
            planned_time=self._parse_datetime(get_val('planned_time')),
            last_sent_date=self._parse_date(get_val('last_sent_date')),
            sent_time=self._parse_datetime(get_val('sent_time')),
            variables=variables,
        )

    def _parse_datetime(self, value) -> Optional[datetime]:
        """解析日期时间值"""
        if value is None:
            return None
        if isinstance(value, datetime):
            return value
        if isinstance(value, str) and value.strip():
            try:
                return datetime.fromisoformat(value.strip())
            except ValueError:
                return None
        return None

    def _parse_date(self, value) -> Optional[date]:
        """解析日期值"""
        if value is None:
            return None
        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, date):
            return value
        if isinstance(value, str) and value.strip():
            try:
                return date.fromisoformat(value.strip())
            except ValueError:
                return None
        return None
