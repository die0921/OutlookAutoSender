import openpyxl
from datetime import datetime, date
from typing import Optional
from src.models.config import ExcelConfig


class ExcelWriter:
    """更新 Excel 文件中的邮件任务状态"""

    def __init__(self, config: ExcelConfig):
        self.config = config

    def update_status(self, file_path: str, row_index: int, status: str) -> None:
        """更新指定行的发送状态"""
        wb = openpyxl.load_workbook(file_path)
        ws = wb[self.config.sheet_name]
        headers = self._get_headers(ws)

        col_name = self.config.columns.get('status', '发送状态')
        col_idx = headers.get(col_name)
        if col_idx:
            ws.cell(row=row_index, column=col_idx, value=status)

        wb.save(file_path)
        wb.close()

    def update_sent(self, file_path: str, row_index: int,
                    status: str, sent_time: datetime, last_sent_date: date) -> None:
        """发送成功后更新状态、实际发送时间、最后发送日期"""
        wb = openpyxl.load_workbook(file_path)
        ws = wb[self.config.sheet_name]
        headers = self._get_headers(ws)

        def set_col(col_key: str, value) -> None:
            col_name = self.config.columns.get(col_key, '')
            col_idx = headers.get(col_name)
            if col_idx and value is not None:
                ws.cell(row=row_index, column=col_idx, value=value)

        set_col('status', status)
        set_col('sent_time', sent_time)
        set_col('last_sent_date', last_sent_date)

        wb.save(file_path)
        wb.close()

    def update_planned_time(self, file_path: str, row_index: int,
                            planned_time: Optional[datetime]) -> None:
        """更新计划发送时间"""
        wb = openpyxl.load_workbook(file_path)
        ws = wb[self.config.sheet_name]
        headers = self._get_headers(ws)

        col_name = self.config.columns.get('planned_time', '计划发送时间')
        col_idx = headers.get(col_name)
        if col_idx:
            ws.cell(row=row_index, column=col_idx, value=planned_time)

        wb.save(file_path)
        wb.close()

    def clear_planned_time(self, file_path: str, row_index: int) -> None:
        """清空计划发送时间（重复任务重置时）"""
        self.update_planned_time(file_path, row_index, None)

    def reset_to_pending(self, file_path: str, row_index: int) -> None:
        """将任务重置为待发送状态，并清空计划发送时间"""
        wb = openpyxl.load_workbook(file_path)
        ws = wb[self.config.sheet_name]
        headers = self._get_headers(ws)

        status_col = self.config.columns.get('status', '发送状态')
        planned_col = self.config.columns.get('planned_time', '计划发送时间')

        if headers.get(status_col):
            ws.cell(row=row_index, column=headers[status_col], value='待发送')
        if headers.get(planned_col):
            cell = ws.cell(row=row_index, column=headers[planned_col])
            cell.value = None

        wb.save(file_path)
        wb.close()

    def _get_headers(self, ws) -> dict:
        """读取表头行，返回列名到列索引的映射"""
        headers = {}
        for col_idx, cell in enumerate(ws[1], start=1):
            if cell.value:
                headers[str(cell.value)] = col_idx
        return headers
