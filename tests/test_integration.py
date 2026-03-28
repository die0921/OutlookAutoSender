"""
Integration tests for OutlookAutoSender end-to-end flows.

Each test uses:
  - Real temporary Excel files (openpyxl)
  - Real ExcelReader / ExcelWriter with the temp file
  - Real ScheduleCalculator / TemplateEngine (where applicable)
  - Mocked Outlook COM calls (EmailService.send_email or _send_via_outlook)
"""
import os
import tempfile
from datetime import datetime, date, timedelta
from typing import Optional
from unittest.mock import MagicMock, patch

import openpyxl
import pytest

from src.data.excel_reader import ExcelReader
from src.data.excel_writer import ExcelWriter
from src.models.config import (
    AppConfig,
    Config,
    ErrorHandlingConfig,
    ExcelConfig,
    LoggingConfig,
    OutlookConfig,
    WorkdaysConfig,
)
from src.models.template import BodyPart, Template
from src.services.email_service import EmailService
from src.services.outlook_contact_service import OutlookContactService
from src.services.schedule_calculator import ScheduleCalculator
from src.services.scheduler_service import SchedulerService
from src.services.template_engine import TemplateEngine


# ---------------------------------------------------------------------------
# Column configuration shared by all tests
# ---------------------------------------------------------------------------
COLUMNS = {
    "recipient": "收件人",
    "cc": "抄送",
    "subject": "主题",
    "attachments": "附件",
    "status": "发送状态",
    "send_mode": "发送模式",
    "template": "模板名称",
    "schedule_type": "发送方式",
    "schedule_params": "发送参数",
    "planned_time": "计划发送时间",
    "last_sent_date": "最后发送日期",
    "sent_time": "实际发送时间",
}

HEADERS = [
    "收件人",
    "抄送",
    "主题",
    "附件",
    "发送状态",
    "发送模式",
    "模板名称",
    "发送方式",
    "发送参数",
    "计划发送时间",
    "最后发送日期",
    "实际发送时间",
]


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def make_config(excel_file: str) -> Config:
    return Config(
        app=AppConfig(name="Test", check_interval=60, expiry_notice_minutes=30),
        excel=ExcelConfig(
            main_file=excel_file,
            sheet_name="Sheet1",
            columns=COLUMNS,
        ),
        outlook=OutlookConfig(account="", save_to_sent=False),
        workdays=WorkdaysConfig(weekdays=[1, 2, 3, 4, 5], holidays=[]),
        error_handling=ErrorHandlingConfig(
            stop_scheduler=True,
            skip_and_continue=False,
            mark_failed_as="发送失败",
            show_dialog=False,
            sound_alert=False,
        ),
        logging=LoggingConfig(
            level="INFO", file="logs/app.log", max_size_mb=10, backup_count=5
        ),
        status_values={"pending": "待发送", "sent": "已发送", "failed": "发送失败"},
        send_mode_values={"repeat": "重复", "once": "一次性"},
        templates_file="templates.yaml",
    )


def _col_index(header_name: str) -> int:
    """Return 1-based column index of a header in HEADERS list."""
    return HEADERS.index(header_name) + 1


def create_excel(path: str, rows: list) -> None:
    """Create a test Excel workbook with the shared header layout."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(HEADERS)
    for row in rows:
        ws.append(row)
    wb.save(path)
    wb.close()


def read_cell(path: str, row: int, col_name: str):
    """Read a single cell value from the test Excel file."""
    wb = openpyxl.load_workbook(path)
    ws = wb["Sheet1"]
    col_idx = _col_index(col_name)
    value = ws.cell(row=row, column=col_idx).value
    wb.close()
    return value


def _empty_row() -> list:
    """Return a list of None values matching HEADERS length."""
    return [None] * len(HEADERS)


def make_row(
    recipient: str = "user@example.com",
    subject: str = "Test Subject",
    status: str = "待发送",
    send_mode: str = "一次性",
    schedule_type: str = "按天",
    schedule_params: str = "09:00",
    planned_time=None,
    last_sent_date=None,
    template_name: str = "",
    cc: str = "",
) -> list:
    row = _empty_row()
    row[HEADERS.index("收件人")] = recipient
    row[HEADERS.index("抄送")] = cc
    row[HEADERS.index("主题")] = subject
    row[HEADERS.index("发送状态")] = status
    row[HEADERS.index("发送模式")] = send_mode
    row[HEADERS.index("发送方式")] = schedule_type
    row[HEADERS.index("发送参数")] = schedule_params
    row[HEADERS.index("计划发送时间")] = planned_time
    row[HEADERS.index("最后发送日期")] = last_sent_date
    row[HEADERS.index("模板名称")] = template_name
    return row


# ---------------------------------------------------------------------------
# Scenario 1: 按天重复发送端到端流程
#   - Real Excel file with one "按天" task in "待发送" state, planned_time in past
#   - Mock EmailService.send_email to return (True, "")
#   - Verify Excel row is updated to "已发送"
# ---------------------------------------------------------------------------

class TestDailyRepeatEndToEnd:
    def test_daily_repeat_send_success(self, tmp_path):
        """按天重复任务到期 → 发送 → Excel 状态更新为已发送"""
        excel_path = str(tmp_path / "tasks.xlsx")
        past_time = datetime.now() - timedelta(minutes=10)

        create_excel(
            excel_path,
            [
                make_row(
                    recipient="alice@example.com",
                    subject="Daily Report",
                    status="待发送",
                    send_mode="重复",
                    schedule_type="按天",
                    schedule_params="09:00",
                    planned_time=past_time,
                )
            ],
        )

        config = make_config(excel_path)
        excel_reader = ExcelReader(config.excel)
        excel_writer = ExcelWriter(config.excel)
        schedule_calc = ScheduleCalculator(
            workday_numbers=config.workdays.weekdays,
            holidays=config.workdays.holidays,
        )

        mock_email_service = MagicMock(spec=EmailService)
        mock_email_service.send_email.return_value = (True, "")

        scheduler = SchedulerService(
            config=config,
            excel_reader=excel_reader,
            excel_writer=excel_writer,
            email_service=mock_email_service,
            schedule_calculator=schedule_calc,
        )

        scheduler.start()

        # Verify send was called
        mock_email_service.send_email.assert_called_once()
        sent_task = mock_email_service.send_email.call_args[0][0]
        assert sent_task.recipient == "alice@example.com"

        # Verify Excel was updated to "已发送"
        status_in_file = read_cell(excel_path, row=2, col_name="发送状态")
        assert status_in_file == "已发送", (
            f"Expected '已发送' but got '{status_in_file}'"
        )

    def test_daily_repeat_assigns_planned_time(self, tmp_path):
        """按天重复任务无计划时间 → 调度器写入计划发送时间"""
        excel_path = str(tmp_path / "tasks.xlsx")

        create_excel(
            excel_path,
            [
                make_row(
                    recipient="bob@example.com",
                    status="待发送",
                    send_mode="重复",
                    schedule_type="按天",
                    schedule_params="23:59",  # far in future today
                    planned_time=None,
                )
            ],
        )

        config = make_config(excel_path)
        excel_reader = ExcelReader(config.excel)
        excel_writer = ExcelWriter(config.excel)
        schedule_calc = ScheduleCalculator(
            workday_numbers=config.workdays.weekdays,
            holidays=config.workdays.holidays,
        )

        mock_email_service = MagicMock(spec=EmailService)
        mock_email_service.send_email.return_value = (True, "")

        scheduler = SchedulerService(
            config=config,
            excel_reader=excel_reader,
            excel_writer=excel_writer,
            email_service=mock_email_service,
            schedule_calculator=schedule_calc,
        )

        scheduler.start()

        planned = read_cell(excel_path, row=2, col_name="计划发送时间")
        assert planned is not None, "Planned time should have been written to Excel"


# ---------------------------------------------------------------------------
# Scenario 2: 按日期列表任务端到端
#   - Task with schedule_type "按日期列表" and a date in the past
#   - planned_time already set to a past datetime
#   - Verify the task is sent and Excel updated
# ---------------------------------------------------------------------------

class TestDateListEndToEnd:
    def test_date_list_due_task_sent(self, tmp_path):
        """按日期列表任务到期 → 发送 → Excel 更新"""
        excel_path = str(tmp_path / "tasks.xlsx")

        # Planned time is 10 minutes in the past
        past_time = datetime.now() - timedelta(minutes=10)

        create_excel(
            excel_path,
            [
                make_row(
                    recipient="carol@example.com",
                    subject="Date List Task",
                    status="待发送",
                    send_mode="一次性",
                    schedule_type="按日期列表",
                    schedule_params="2020-01-01,2020-06-15;09:00",
                    planned_time=past_time,
                )
            ],
        )

        config = make_config(excel_path)
        excel_reader = ExcelReader(config.excel)
        excel_writer = ExcelWriter(config.excel)
        schedule_calc = ScheduleCalculator(
            workday_numbers=config.workdays.weekdays,
            holidays=config.workdays.holidays,
        )

        mock_email_service = MagicMock(spec=EmailService)
        mock_email_service.send_email.return_value = (True, "")

        scheduler = SchedulerService(
            config=config,
            excel_reader=excel_reader,
            excel_writer=excel_writer,
            email_service=mock_email_service,
            schedule_calculator=schedule_calc,
        )

        scheduler.start()

        mock_email_service.send_email.assert_called_once()
        status_in_file = read_cell(excel_path, row=2, col_name="发送状态")
        assert status_in_file == "已发送"

    def test_date_list_future_task_not_sent(self, tmp_path):
        """按日期列表任务未到期 → 不发送，状态保持待发送"""
        excel_path = str(tmp_path / "tasks.xlsx")

        # planned_time is far in the future (no due tasks)
        future_time = datetime.now() + timedelta(days=30)

        create_excel(
            excel_path,
            [
                make_row(
                    recipient="dave@example.com",
                    status="待发送",
                    send_mode="一次性",
                    schedule_type="按日期列表",
                    schedule_params="2099-12-31;09:00",
                    planned_time=future_time,
                )
            ],
        )

        config = make_config(excel_path)
        excel_reader = ExcelReader(config.excel)
        excel_writer = ExcelWriter(config.excel)
        schedule_calc = ScheduleCalculator(
            workday_numbers=config.workdays.weekdays,
            holidays=config.workdays.holidays,
        )

        mock_email_service = MagicMock(spec=EmailService)

        scheduler = SchedulerService(
            config=config,
            excel_reader=excel_reader,
            excel_writer=excel_writer,
            email_service=mock_email_service,
            schedule_calculator=schedule_calc,
        )

        scheduler.start()

        mock_email_service.send_email.assert_not_called()
        status_in_file = read_cell(excel_path, row=2, col_name="发送状态")
        assert status_in_file == "待发送"


# ---------------------------------------------------------------------------
# Scenario 3: 模板渲染完整流程
#   - EmailService with real TemplateEngine (mocked TemplateLoader)
#   - Task references a template; send_email should render subject/body
#   - Mock _send_via_outlook to avoid COM dependency
# ---------------------------------------------------------------------------

class TestTemplateRenderEndToEnd:
    def test_template_rendered_before_send(self, tmp_path):
        """模板渲染完整流程：变量替换后正确渲染主题和正文"""
        excel_path = str(tmp_path / "tasks.xlsx")
        past_time = datetime.now() - timedelta(minutes=5)

        create_excel(
            excel_path,
            [
                make_row(
                    recipient="eve@example.com",
                    status="待发送",
                    send_mode="一次性",
                    schedule_type="按天",
                    schedule_params="09:00",
                    planned_time=past_time,
                    template_name="monthly_report",
                )
            ],
        )

        # Build a real TemplateEngine with a mocked TemplateLoader
        mock_loader = MagicMock()
        template = Template(
            name="monthly_report",
            subject="月报 - {{month}}",
            body_type="text",
            body_parts=[BodyPart(type="text", content="你好 {{name}}，本月报告已就绪。")],
            data_sources={},
        )
        mock_loader.load.return_value = template
        template_engine = TemplateEngine(template_loader=mock_loader)

        # Contact service: return the email address directly
        mock_contact_service = MagicMock(spec=OutlookContactService)
        mock_contact_service.resolve_recipients.side_effect = lambda s: [s]

        rendered_calls = []

        # Patch _send_via_outlook to capture what was sent
        def fake_send(recipients, cc_list, subject, body, attachments):
            rendered_calls.append({"subject": subject, "body": body})

        email_service = EmailService(
            template_engine=template_engine,
            contact_service=mock_contact_service,
        )
        email_service._send_via_outlook = fake_send

        config = make_config(excel_path)
        excel_reader = ExcelReader(config.excel)
        excel_writer = ExcelWriter(config.excel)
        schedule_calc = ScheduleCalculator(
            workday_numbers=config.workdays.weekdays,
            holidays=config.workdays.holidays,
        )

        scheduler = SchedulerService(
            config=config,
            excel_reader=excel_reader,
            excel_writer=excel_writer,
            email_service=email_service,
            schedule_calculator=schedule_calc,
        )

        scheduler.start()

        # The template was loaded and rendered
        mock_loader.load.assert_called_with("monthly_report")
        assert len(rendered_calls) == 1, "Expected exactly one send call"

        # Variables dict is empty (no extra columns in test Excel), so
        # Jinja2 renders undefined vars as empty string via Undefined class
        # Subject should still contain "月报 - " prefix
        assert "月报" in rendered_calls[0]["subject"]

        # Excel status updated
        status_in_file = read_cell(excel_path, row=2, col_name="发送状态")
        assert status_in_file == "已发送"

    def test_template_with_variables(self, tmp_path):
        """模板变量替换：Excel 列数据正确注入模板"""
        excel_path = str(tmp_path / "tasks.xlsx")
        past_time = datetime.now() - timedelta(minutes=5)

        # Add extra columns (姓名, 月份) as template variables
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        extended_headers = HEADERS + ["姓名", "月份"]
        ws.append(extended_headers)

        row_data = [None] * len(extended_headers)
        row_data[extended_headers.index("收件人")] = "frank@example.com"
        row_data[extended_headers.index("发送状态")] = "待发送"
        row_data[extended_headers.index("发送模式")] = "一次性"
        row_data[extended_headers.index("发送方式")] = "按天"
        row_data[extended_headers.index("发送参数")] = "09:00"
        row_data[extended_headers.index("计划发送时间")] = past_time
        row_data[extended_headers.index("模板名称")] = "greeting"
        row_data[extended_headers.index("姓名")] = "Frank"
        row_data[extended_headers.index("月份")] = "三月"
        ws.append(row_data)
        wb.save(excel_path)
        wb.close()

        mock_loader = MagicMock()
        template = Template(
            name="greeting",
            subject="{{月份}}月报",
            body_type="text",
            body_parts=[BodyPart(type="text", content="你好 {{姓名}}！")],
            data_sources={},
        )
        mock_loader.load.return_value = template
        template_engine = TemplateEngine(template_loader=mock_loader)

        mock_contact_service = MagicMock(spec=OutlookContactService)
        mock_contact_service.resolve_recipients.side_effect = lambda s: [s]

        rendered_calls = []

        def fake_send(recipients, cc_list, subject, body, attachments):
            rendered_calls.append({"subject": subject, "body": body})

        email_service = EmailService(
            template_engine=template_engine,
            contact_service=mock_contact_service,
        )
        email_service._send_via_outlook = fake_send

        config = make_config(excel_path)
        excel_reader = ExcelReader(config.excel)
        excel_writer = ExcelWriter(config.excel)
        schedule_calc = ScheduleCalculator(
            workday_numbers=config.workdays.weekdays,
            holidays=config.workdays.holidays,
        )

        scheduler = SchedulerService(
            config=config,
            excel_reader=excel_reader,
            excel_writer=excel_writer,
            email_service=email_service,
            schedule_calculator=schedule_calc,
        )

        scheduler.start()

        assert len(rendered_calls) == 1
        assert "三月" in rendered_calls[0]["subject"]
        assert "Frank" in rendered_calls[0]["body"]


# ---------------------------------------------------------------------------
# Scenario 4: 发送失败 → 停止调度器
#   - Task with planned_time in past
#   - Mock EmailService.send_email returns (False, "Connection failed")
#   - Verify scheduler stops, on_error is called, Excel updated to "发送失败"
# ---------------------------------------------------------------------------

class TestFailureStopsScheduler:
    def test_send_failure_stops_scheduler_and_updates_excel(self, tmp_path):
        """发送失败 → 调度器停止 + Excel 状态更新为发送失败"""
        excel_path = str(tmp_path / "tasks.xlsx")
        past_time = datetime.now() - timedelta(minutes=10)

        create_excel(
            excel_path,
            [
                make_row(
                    recipient="grace@example.com",
                    subject="Will Fail",
                    status="待发送",
                    send_mode="一次性",
                    schedule_type="按天",
                    schedule_params="09:00",
                    planned_time=past_time,
                )
            ],
        )

        config = make_config(excel_path)
        excel_reader = ExcelReader(config.excel)
        excel_writer = ExcelWriter(config.excel)
        schedule_calc = ScheduleCalculator(
            workday_numbers=config.workdays.weekdays,
            holidays=config.workdays.holidays,
        )

        mock_email_service = MagicMock(spec=EmailService)
        mock_email_service.send_email.return_value = (False, "Connection failed")

        error_callback = MagicMock()

        scheduler = SchedulerService(
            config=config,
            excel_reader=excel_reader,
            excel_writer=excel_writer,
            email_service=mock_email_service,
            schedule_calculator=schedule_calc,
            on_error=error_callback,
        )

        scheduler.start()

        # Scheduler should be stopped
        assert scheduler.get_status().running is False

        # Error message captured
        assert scheduler.get_status().error_message == "Connection failed"

        # on_error callback called with the error message
        error_callback.assert_called_once()
        callback_args = error_callback.call_args[0]
        assert "Connection failed" in callback_args[0]

        # Excel updated to "发送失败"
        status_in_file = read_cell(excel_path, row=2, col_name="发送状态")
        assert status_in_file == "发送失败"

    def test_send_failure_stops_processing_subsequent_tasks(self, tmp_path):
        """第一个任务发送失败 → 后续任务不再发送"""
        excel_path = str(tmp_path / "tasks.xlsx")
        past_time = datetime.now() - timedelta(minutes=10)

        create_excel(
            excel_path,
            [
                make_row(
                    recipient="task1@example.com",
                    subject="Task 1 - Will Fail",
                    status="待发送",
                    send_mode="一次性",
                    schedule_type="按天",
                    schedule_params="09:00",
                    planned_time=past_time,
                ),
                make_row(
                    recipient="task2@example.com",
                    subject="Task 2 - Should Not Send",
                    status="待发送",
                    send_mode="一次性",
                    schedule_type="按天",
                    schedule_params="09:00",
                    planned_time=past_time,
                ),
            ],
        )

        config = make_config(excel_path)
        excel_reader = ExcelReader(config.excel)
        excel_writer = ExcelWriter(config.excel)
        schedule_calc = ScheduleCalculator(
            workday_numbers=config.workdays.weekdays,
            holidays=config.workdays.holidays,
        )

        mock_email_service = MagicMock(spec=EmailService)
        mock_email_service.send_email.return_value = (False, "SMTP error")

        scheduler = SchedulerService(
            config=config,
            excel_reader=excel_reader,
            excel_writer=excel_writer,
            email_service=mock_email_service,
            schedule_calculator=schedule_calc,
        )

        scheduler.start()

        # Only the first task's send should have been attempted
        assert mock_email_service.send_email.call_count == 1

        # Scheduler stopped after first failure
        assert scheduler.get_status().running is False

        # First task marked as failed
        status_row2 = read_cell(excel_path, row=2, col_name="发送状态")
        assert status_row2 == "发送失败"

        # Second task still pending (not attempted)
        status_row3 = read_cell(excel_path, row=3, col_name="发送状态")
        assert status_row3 == "待发送"
