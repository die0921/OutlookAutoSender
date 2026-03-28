import pytest
import os
import openpyxl
from datetime import datetime, date
from src.data.excel_reader import ExcelReader
from src.models.config import ExcelConfig


COLUMNS = {
    'recipient': '收件人',
    'cc': '抄送',
    'subject': '主题',
    'attachments': '附件路径',
    'status': '发送状态',
    'send_mode': '发送模式',
    'template': '使用模板',
    'schedule_type': '发送方式',
    'schedule_params': '发送参数',
    'planned_time': '计划发送时间',
    'last_sent_date': '最后发送日期',
    'sent_time': '实际发送时间',
}


@pytest.fixture
def excel_config():
    return ExcelConfig(
        main_file="",
        sheet_name="邮件列表",
        columns=COLUMNS,
    )


@pytest.fixture
def test_excel_file(tmp_path, excel_config):
    """创建测试 Excel 文件"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = excel_config.sheet_name

    # 写入表头
    headers = list(COLUMNS.values()) + ['客户名称', '部门']
    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx, value=header)

    # 写入数据行
    row1 = {
        '收件人': 'user@example.com',
        '抄送': 'cc@example.com',
        '主题': '测试邮件',
        '附件路径': '',
        '发送状态': '待发送',
        '发送模式': '一次性',
        '使用模板': '简单模板',
        '发送方式': '按天',
        '发送参数': '9:00',
        '计划发送时间': None,
        '最后发送日期': None,
        '实际发送时间': None,
        '客户名称': '张三',
        '部门': '销售部',
    }
    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=2, column=col_idx, value=row1.get(header))

    # 空行（应被跳过）
    ws.cell(row=3, column=1, value=None)

    # 第二条数据
    row3 = {
        '收件人': 'another@example.com',
        '发送状态': '已发送',
        '发送模式': '重复',
        '发送方式': '按周',
        '发送参数': '周一,9:00',
        '客户名称': '李四',
    }
    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=4, column=col_idx, value=row3.get(header))

    path = str(tmp_path / "test_emails.xlsx")
    wb.save(path)
    return path


def test_read_tasks_basic(test_excel_file, excel_config):
    reader = ExcelReader(excel_config)
    tasks = reader.read_tasks(test_excel_file)

    assert len(tasks) == 2
    assert tasks[0].recipient == 'user@example.com'
    assert tasks[0].cc == 'cc@example.com'
    assert tasks[0].status == '待发送'
    assert tasks[0].send_mode == '一次性'


def test_read_tasks_variables(test_excel_file, excel_config):
    reader = ExcelReader(excel_config)
    tasks = reader.read_tasks(test_excel_file)

    assert tasks[0].variables.get('客户名称') == '张三'
    assert tasks[0].variables.get('部门') == '销售部'


def test_read_tasks_skips_empty_rows(test_excel_file, excel_config):
    reader = ExcelReader(excel_config)
    tasks = reader.read_tasks(test_excel_file)

    # Row 3 is empty and should be skipped
    assert len(tasks) == 2
    assert tasks[1].recipient == 'another@example.com'


def test_read_tasks_row_index(test_excel_file, excel_config):
    reader = ExcelReader(excel_config)
    tasks = reader.read_tasks(test_excel_file)

    assert tasks[0].row_index == 2
    assert tasks[1].row_index == 4  # row 3 was empty/skipped
