import pytest
import openpyxl
from datetime import datetime, date
from src.data.excel_writer import ExcelWriter
from src.models.config import ExcelConfig


COLUMNS = {
    'recipient': '收件人',
    'status': '发送状态',
    'send_mode': '发送模式',
    'planned_time': '计划发送时间',
    'last_sent_date': '最后发送日期',
    'sent_time': '实际发送时间',
}


@pytest.fixture
def excel_config():
    return ExcelConfig(main_file="", sheet_name="邮件列表", columns=COLUMNS)


@pytest.fixture
def test_excel_file(tmp_path, excel_config):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = excel_config.sheet_name
    headers = list(COLUMNS.values())
    for col_idx, h in enumerate(headers, 1):
        ws.cell(row=1, column=col_idx, value=h)
    # Data row at row 2
    ws.cell(row=2, column=1, value='user@example.com')
    ws.cell(row=2, column=2, value='待发送')
    path = str(tmp_path / "test.xlsx")
    wb.save(path)
    return path


def _read_cell(file_path, sheet, row, col_name, headers_row=1):
    wb = openpyxl.load_workbook(file_path)
    ws = wb[sheet]
    headers = {str(c.value): i for i, c in enumerate(ws[headers_row], 1) if c.value}
    col_idx = headers.get(col_name)
    val = ws.cell(row=row, column=col_idx).value if col_idx else None
    wb.close()
    return val


def test_update_status(test_excel_file, excel_config):
    writer = ExcelWriter(excel_config)
    writer.update_status(test_excel_file, 2, '已发送')
    assert _read_cell(test_excel_file, '邮件列表', 2, '发送状态') == '已发送'


def test_update_sent(test_excel_file, excel_config):
    writer = ExcelWriter(excel_config)
    sent_time = datetime(2026, 3, 28, 9, 0, 0)
    last_date = date(2026, 3, 28)
    writer.update_sent(test_excel_file, 2, '已发送', sent_time, last_date)

    assert _read_cell(test_excel_file, '邮件列表', 2, '发送状态') == '已发送'
    assert _read_cell(test_excel_file, '邮件列表', 2, '实际发送时间') == sent_time


def test_update_planned_time(test_excel_file, excel_config):
    writer = ExcelWriter(excel_config)
    planned = datetime(2026, 3, 29, 9, 0, 0)
    writer.update_planned_time(test_excel_file, 2, planned)
    assert _read_cell(test_excel_file, '邮件列表', 2, '计划发送时间') == planned


def test_clear_planned_time(test_excel_file, excel_config):
    writer = ExcelWriter(excel_config)
    writer.clear_planned_time(test_excel_file, 2)
    assert _read_cell(test_excel_file, '邮件列表', 2, '计划发送时间') is None


def test_reset_to_pending(test_excel_file, excel_config):
    writer = ExcelWriter(excel_config)
    # First set to sent with planned time
    writer.update_planned_time(test_excel_file, 2, datetime(2026, 3, 28, 9, 0))
    writer.update_status(test_excel_file, 2, '已发送')
    # Now reset
    writer.reset_to_pending(test_excel_file, 2)
    assert _read_cell(test_excel_file, '邮件列表', 2, '发送状态') == '待发送'
    assert _read_cell(test_excel_file, '邮件列表', 2, '计划发送时间') is None
