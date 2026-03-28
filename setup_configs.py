#!/usr/bin/env python3
"""Setup script for config files, templates, and example Excel data."""

import os
import yaml
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# Task 1: Update config.yaml
print("Updating config.yaml...")
config_data = {
    'app': {
        'name': 'Outlook 自动发送助手',
        'check_interval': 60,
        'expiry_notice_minutes': 30
    },
    'excel': {
        'main_file': 'data/emails.xlsx',
        'sheet_name': '邮件列表',
        'columns': {
            'recipient': '收件人',
            'cc': '抄送',
            'subject': '主题',
            'attachments': '附件路径',
            'status': '发送状态',
            'send_mode': '发送模式',
            'template': '使用模板',
            'schedule_type': '发送方式',
            'schedule_params': '发送参数',
            'planned_time': '计划发送时间',
            'last_sent_date': '最后发送日期',
            'sent_time': '实际发送时间'
        }
    },
    'status_values': {
        'pending': '待发送',
        'sent': '已发送',
        'failed': '发送失败',
        'expired': '已过期'
    },
    'send_mode_values': {
        'repeat': '重复',
        'once': '一次性'
    },
    'templates_file': 'templates.yaml',
    'outlook': {
        'account': '',
        'save_to_sent': True
    },
    'workdays': {
        'weekdays': [1, 2, 3, 4, 5],
        'holidays': ['2026-01-01', '2026-10-01']
    },
    'error_handling': {
        'on_error': {
            'stop_scheduler': True,
            'skip_and_continue': False
        },
        'mark_failed_as': '发送失败',
        'notification': {
            'show_dialog': True,
            'sound_alert': True
        }
    },
    'logging': {
        'level': 'INFO',
        'file': 'logs/app.log',
        'max_size_mb': 10,
        'backup_count': 5
    }
}

with open('config.yaml', 'w', encoding='utf-8') as f:
    yaml.dump(config_data, f, allow_unicode=True, default_flow_style=False)
print("✓ config.yaml updated")

# Task 2: Update templates.yaml
print("Updating templates.yaml...")
templates_data = {
    'templates': {
        '日报模板': {
            'subject': '{{ 月份 }}月日报 - {{ 客户名称 }}',
            'body_type': 'html',
            'body_parts': [
                {
                    'type': 'text',
                    'content': '尊敬的 {{ 客户名称 }}：\n\n以下是您的日报内容。\n\n此邮件由系统自动发送。'
                }
            ]
        },
        '月报模板': {
            'subject': '{{ 月份 }}月度报告 - {{ 客户名称 }}',
            'body_type': 'html',
            'body_parts': [
                {
                    'type': 'text',
                    'content': '尊敬的 {{ 客户名称 }}：\n\n请查收本月报告。\n\n此邮件由系统自动发送。'
                }
            ]
        },
        '简单模板': {
            'subject': '{{ 主题 }}',
            'body_type': 'text',
            'body_parts': [
                {
                    'type': 'text',
                    'content': '{{ 正文 }}'
                }
            ]
        }
    },
    'global_attachments': []
}

with open('templates.yaml', 'w', encoding='utf-8') as f:
    yaml.dump(templates_data, f, allow_unicode=True, default_flow_style=False)
print("✓ templates.yaml updated")

# Task 3: Create data/ directory and data/emails.xlsx
print("Creating data/ directory and emails.xlsx...")
os.makedirs('data', exist_ok=True)

# Create Excel file
wb = Workbook()
ws = wb.active
ws.title = '邮件列表'

# Headers
headers = ['收件人', '抄送', '主题', '附件路径', '发送状态', '发送模式', '使用模板',
           '发送方式', '发送参数', '计划发送时间', '最后发送日期', '实际发送时间', '客户名称']

# Style for headers
header_fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
header_font = Font(bold=True)

for col_idx, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_idx)
    cell.value = header
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center')

# Example data
example_rows = [
    ['user@example.com', '', '日报', '', '待发送', '重复', '', '按天', '9:00', '', '', '', '张三'],
    ['manager@example.com', '', '月报', '', '待发送', '一次性', '', '按月-固定', '1日,9:00', '', '', '', '李四'],
    ['team@example.com', '', '周报', '', '待发送', '重复', '', '按周', '周一,9:00', '', '', '', '王五']
]

for row_idx, row_data in enumerate(example_rows, 2):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.value = value
        cell.alignment = Alignment(horizontal='left')

# Adjust column widths
col_widths = [20, 20, 15, 20, 12, 12, 15, 15, 15, 20, 20, 20, 12]
for col_idx, width in enumerate(col_widths, 1):
    col_letter = chr(64 + col_idx)
    ws.column_dimensions[col_letter].width = width

wb.save('data/emails.xlsx')
print("✓ data/emails.xlsx created")

print("\nAll files updated successfully!")
